# -*- coding: utf-8 -*-
"""
Created on Mon Jan 07 11:23:11 2019

@author: woute
"""

from math import *
import csv
import openpyxl as xl
import numpy as np
import gurobipy as grb

from Excel_data_P3 import *
from Functions_P3 import *


wb = xl.load_workbook("outP3 - KPI.xlsx", read_only=True)
ws = wb['outP3 - it3']


direct   = np.zeros(shape=(24,24))
transfer = np.zeros(shape=(24,24))
pax      = np.zeros(shape=(24,24))
flights  = np.zeros(shape=(24,24,24))

direct_ex   = np.array([[str(i.value) for i in j] for j in ws['A1':'B90']])
transfer_ex = np.array([[str(i.value) for i in j] for j in ws['D1':'E115']])
flights_ex =  np.array([[str(i.value) for i in j] for j in ws['G1':'H126']])

for c in range(len(direct_ex)):
    for i in range(24):
        for j in range(24):
            for w in range(2):
                if i < 10 and j< 10:
                    if direct_ex[c][0][2] == str(i) and direct_ex[c][0][4] == str(j) and direct_ex[c][0][6] == str(w):
                        direct[i][j] = direct[i][j] + int(direct_ex[c][1])
                        
                elif i >= 10 and j < 10:
                    if direct_ex[c][0][2]+direct_ex[c][0][3] == str(i) and direct_ex[c][0][5] == str(j) and direct_ex[c][0][7] == str(w):
                        direct[i][j] = direct[i][j] + int(direct_ex[c][1])
                        
                elif i < 10 and j >= 10:
                    if direct_ex[c][0][2] == str(i) and direct_ex[c][0][4]+direct_ex[c][0][5] == str(j) and direct_ex[c][0][7] == str(w):
                        direct[i][j] = direct[i][j] + int(direct_ex[c][1])
                        
                else:
                    if direct_ex[c][0][2]+direct_ex[c][0][3] == str(i) and direct_ex[c][0][5]+direct_ex[c][0][6] == str(j) and direct_ex[c][0][8] == str(w):
                        direct[i][j] = direct[i][j] + int(direct_ex[c][1])
                        

for c in range(len(transfer_ex)):
    for i in range(24):
        for j in range(24):
            for w in range(2):
                if i < 10 and j< 10:
                    if transfer_ex[c][0][2] == str(i) and transfer_ex[c][0][4] == str(j) and transfer_ex[c][0][6] == str(w):
                        transfer[i][j] = transfer[i][j] + int(transfer_ex[c][1])
                        
                elif i >= 10 and j < 10:
                    if transfer_ex[c][0][2]+transfer_ex[c][0][3] == str(i) and transfer_ex[c][0][5] == str(j) and transfer_ex[c][0][7] == str(w):
                        transfer[i][j] = transfer[i][j] + int(transfer_ex[c][1])
                        
                elif i < 10 and j >= 10:
                    if transfer_ex[c][0][2] == str(i) and transfer_ex[c][0][4]+transfer_ex[c][0][5] == str(j) and transfer_ex[c][0][7] == str(w):
                        transfer[i][j] = transfer[i][j] + int(transfer_ex[c][1])
                        
                else:
                    if transfer_ex[c][0][2]+transfer_ex[c][0][3] == str(i) and transfer_ex[c][0][5]+transfer_ex[c][0][6] == str(j) and transfer_ex[c][0][8] == str(w):
                        transfer[i][j] = transfer[i][j] + int(transfer_ex[c][1])


for c in range(len(flights_ex)):
    for i in range(24):
        for j in range(24):
            for k in range(5):
                if i < 10 and j< 10:
                    if flights_ex[c][0][2] == str(i) and flights_ex[c][0][4] == str(j) and flights_ex[c][0][6] == str(k):
                        flights[i][j][k] = flights[i][j][k] + int(flights_ex[c][1])
                        
                elif i >= 10 and j < 10:
                    if flights_ex[c][0][2]+flights_ex[c][0][3] == str(i) and flights_ex[c][0][5] == str(j) and flights_ex[c][0][7] == str(k):
                        flights[i][j][k] = flights[i][j][k] + int(flights_ex[c][1])
                        
                elif i < 10 and j >= 10:
                    if flights_ex[c][0][2] == str(i) and flights_ex[c][0][4]+flights_ex[c][0][5] == str(j) and flights_ex[c][0][7] == str(k):
                        flights[i][j][k] = flights[i][j][k] + int(flights_ex[c][1])
                        
                else:
                    if flights_ex[c][0][2]+flights_ex[c][0][3] == str(i) and flights_ex[c][0][5]+flights_ex[c][0][6] == str(j) and flights_ex[c][0][8] == str(k):
                        flights[i][j][k] = flights[i][j][k] + int(flights_ex[c][1])
                        
                        
transfer_pax = np.zeros(shape=(24,24))
for i in range(24):
    for j in range(24):
        if transfer[i][j] > 0.:
            transfer_pax[i][0] = transfer_pax[i][0] + transfer[i][j]
            transfer_pax[0][j] = transfer_pax[0][j] + transfer[i][j]
        

for i in range(24):
    for j in range(24):
        pax[i][j] = direct[i][j] + transfer_pax[i][j]    





KPI = np.array(['From','To', 'Distance', 'Seats', 'Passengers', '', 'ASK', 'RPK','', 'cost', 'revenue', 'RASK', 'CASK', 'Yield'])
KPI_values = []


tot_dist = 0.
tot_seats = 0.
tot_pax = 0.
tot_ASK = 0.
tot_RPK = 0.
tot_cost = 0.
tot_rev = 0.


for i in range(24):
    for j in range(24):
        if pax[i][j] > 0.:
            seats_available = grb.LinExpr.getValue(grb.quicksum(flights[i][j][k] * seats[0][k] for k in range(5)))
            
            ASK = distance(i,j) * seats_available
            RPK = distance(i,j) * pax[i][j]
            
            revenue = y(i,j)*distance(i,j)*pax[i][j]
            cost = grb.LinExpr.getValue(grb.quicksum(flights[i][j][k]*cost_fact(i,j)*(Cx_ar[0][k]+Ct[0][k]*distance(i,j)/speed[0][k]+Cf[0][k]*F22*distance(i,j)/1.5) for k in range(5)))
            
            RASK  = revenue/ASK
            CASK  = cost/ASK
            Yield = revenue/RPK

            KPI_values = np.array([i,j,distance(i,j),seats_available, pax[i][j], '', ASK, RPK, '', cost, revenue, RASK, CASK, Yield])
            KPI = np.vstack( (KPI, KPI_values))
            

            tot_dist = tot_dist + distance(i,j)
            tot_seats = tot_seats + seats_available
            tot_pax = tot_pax + pax[i][j]
            tot_cost = tot_cost + cost
            tot_rev = tot_rev + revenue
            tot_ASK = tot_ASK + ASK
            tot_RPK = tot_RPK + RPK


KPI_tot = np.array(['TOTAL', '', tot_dist, tot_seats, tot_pax, '', tot_ASK, tot_RPK, '', tot_cost, tot_rev, tot_rev/tot_ASK, tot_cost/tot_ASK, tot_rev/tot_RPK])        
KPI = np.vstack( (KPI, KPI_tot))

print tot_rev-tot_cost


with open('KPI_P3_results_actpax.csv', 'wb') as myfile:
     wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
     wr.writerows(KPI)



























                    
                                               